#!/usr/bin/env python3
"""
import_vendor_xlsx.py — Parse a Vendor Payment Data xlsx and import invoice
records into the finance-platform vendor_invoices table via the API.

Expected xlsx schema (row 1 = headers):
  S.no | Supplier Name | GST Number | Contract NO | Date | Systam Invoice No
  | Type | KPI Application number | Status | Invoice Application Number
  | Payment Application Number | Status

Invoice-row detection — a row is an invoice when:
  * Column F (Systam Invoice No) is non-empty AND
  * Column F does NOT contain any of: total / oppo / profit / rate / fm+ot
    (these mark the FM+OT/P/RMB profit-calc blocks below the invoice table)

Fields mapped to vendor_invoices:
  invoice_no      <- F (Systam Invoice No)
  vendor          <- B (Supplier Name) — short form (first token)
  vendor_full     <- B (Supplier Name) — full trimmed string
  invoice_date    <- E (Date) -> YYYY-MM-DD
  period          <- derived from E as YYYY-MM
  site            <- parsed from G (e.g. "Noida")
  service         <- G (Type) verbatim
  bpm_no          <- K (Payment Application Number, HTFK...)
  note            <- D (Contract NO)
  ex_tax/gst/amount <- 0 (xlsx has no per-invoice amounts)
  pay_status      <- "未付" (conservative; Judy marks paid)
  src             <- xlsx path

Usage:
  python import_vendor_xlsx.py --xlsx PATH [--base-url URL] \
         --username USER --password PASS [--dry-run]
"""
import argparse
import http.cookiejar
import json
import re
import sys
import urllib.error
import urllib.request
from datetime import date, datetime

import openpyxl

CALC_KEYWORDS = ("total", "oppo", "profit", "rate", "fm+ot")

# When splitting vendor into short/full, treat these as the 2nd-token discriminator.
VENDOR_SPLIT_TOKENS = {"fireman", "hk", "material", "ind", "industries"}


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def to_date(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def is_invoice_row(f_val):
    if f_val is None or str(f_val).strip() == "":
        return False
    s = str(f_val).strip().lower()
    return not any(k in s for k in CALC_KEYWORDS)


def split_vendor(supplier):
    s = supplier.strip()
    tokens = s.split()
    if len(tokens) >= 2 and (tokens[1].lower() in VENDOR_SPLIT_TOKENS or tokens[1].isupper()):
        return tokens[0], s
    return s, s


def parse_site(type_str):
    if not type_str:
        return ""
    t = re.sub(r"^\d+\)\s*", "", type_str.strip())
    parts = t.split()
    return parts[0] if parts else ""


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    invoices = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [cell(c.value) for c in ws[1]]
        idx = {h.lower(): i for i, h in enumerate(header)}

        def col(*names, default=-1):
            for n in names:
                if n.lower() in idx:
                    return idx[n.lower()]
            return default

        cB = col("Supplier Name")
        cD = col("Contract NO")
        cE = col("Date")
        cF = col("Systam Invoice No")
        cG = col("Type")
        cK = col("Payment Application Number")
        if cF < 0 or cB < 0:
            continue
        for row in ws.iter_rows(min_row=2, values_only=False):
            if cF >= len(row):
                continue
            f_val = row[cF].value
            if not is_invoice_row(f_val):
                continue
            if cB >= len(row):
                continue
            supplier = cell(row[cB].value)
            if not supplier:
                continue
            inv_no = cell(f_val)
            date_iso = to_date(row[cE].value) if cE < len(row) else ""
            type_str = cell(row[cG].value) if cG < len(row) else ""
            bpm = cell(row[cK].value) if cK < len(row) else ""
            contract = cell(row[cD].value) if cD < len(row) else ""
            vshort, vfull = split_vendor(supplier)
            invoices.append({
                "invoice_no": inv_no,
                "vendor": vshort,
                "vendor_full": vfull,
                "invoice_date": date_iso,
                "period": (date_iso[:7] if date_iso else ""),
                "site": parse_site(type_str),
                "service": type_str,
                "ex_tax": 0,
                "gst": 0,
                "amount": 0,
                "currency": "INR",
                "src": str(path),
                "pay_status": "未付",
                "bpm_no": bpm,
                "note": contract,
            })
    return invoices


def main():
    ap = argparse.ArgumentParser(description="Import Vendor Payment Data xlsx into finance-platform via API.")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--base-url", default="http://127.0.0.1:8137")
    ap.add_argument("--username", required=True)
    ap.add_argument("--password", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    invoices = parse_xlsx(args.xlsx)
    print(f"[parse] {len(invoices)} invoice rows from {args.xlsx}")
    for inv in invoices:
        print(
            f"  {inv['period']} | {inv['vendor']:<14} | {inv['invoice_no']:<40} | "
            f"{inv['invoice_date']} | bpm={inv['bpm_no']:<22} | site={inv['site']:<8} | service={inv['service']}"
        )

    if args.dry_run or not invoices:
        return

    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    # login
    data = json.dumps({"username": args.username, "password": args.password}).encode()
    req = urllib.request.Request(args.base_url + "/api/login", data=data, method="POST")
    req.add_header("Content-Type", "application/json")
    try:
        resp = opener.open(req, timeout=30)
    except urllib.error.HTTPError as e:
        print(f"[login] FAILED HTTP {e.code}: {e.read().decode()}")
        sys.exit(1)
    login_body = resp.read().decode()
    if resp.status != 200:
        print(f"[login] FAILED HTTP {resp.status}: {login_body}")
        sys.exit(1)
    print(f"[login] ok as {args.username}")

    # import
    req = urllib.request.Request(
        args.base_url + "/api/vendor-invoice-import",
        data=json.dumps({"rows": invoices}).encode(),
        method="POST",
    )
    req.add_header("Content-Type", "application/json")
    try:
        resp = opener.open(req, timeout=60)
    except urllib.error.HTTPError as e:
        print(f"[import] FAILED HTTP {e.code}: {e.read().decode()}")
        sys.exit(1)
    body = resp.read().decode()
    print(f"[import] HTTP {resp.status}")
    try:
        print(json.dumps(json.loads(body), ensure_ascii=False, indent=2))
    except Exception:
        print(body)


if __name__ == "__main__":
    main()